from django.shortcuts import render,redirect
from bloodDonationApp.models import *
# import datetime
from datetime import datetime,timedelta,date
from django.contrib import messages
from django.http import JsonResponse,HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,Alignment,Border,Side
from django.utils import timezone


# Create your views here.

def admin_home(request): 
    if 'uid' not in request.session:
        return redirect('/login')
    else:
        return render(request,'admin.html')
def add_organization(request):
    if 'uid' not in request.session:
        return redirect('/login')
    else:
        return render(request,'add_organization.html')
def add_organization_db(request):
    if 'uid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            org1 = organizationMaster()
            org1.Organization_name = request.POST.get('organizationName')
            org1.Address = request.POST.get('address')
            org1.Phone_no = request.POST.get('phone')
            org1.Email = request.POST.get('email')
            org1.Status = "YES"
            org1.Created_date = date.today().strftime("%Y-%m-%d")
            org1.save()
            messages.success(request,"Organization added successfully")
            return redirect('/add_organization')
def delete_organaization(request):
    if 'uid' not in request.session:
        return redirect('/login')
    else:
        orgs = organizationMaster.objects.all()
        return render(request,'delete_organization.html',{'orgs':orgs})
def delete_organization_db(request,Ornganization_id):
    if 'uid' not in request.session:
        return redirect('/login')
    else:
        data = organizationMaster.objects.get(Ornganization_id=Ornganization_id)
        data.delete()
        messages.success(request,"Organization removed successfully")
        return redirect('/delete_organaization')
def add_superUser(request):
    if 'uid' not in request.session:
        return redirect('/login')
    else:
        data = organizationMaster.objects.all()
        return render(request,'Add_superuser.html',{'data':data})
def add_superUser_db(request):
    if 'uid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            Ornganization_id = request.POST.get('organization')
            existing_user = userMaster.objects.filter(Ornganization_id_id = Ornganization_id, Role="SU").count()
            # print(existing_user)
            if existing_user > 0:
                messages.error(request, "This organization already has a SuperUser.")
                return redirect('/add_superUser')
            um = userMaster()
            um.Username = request.POST.get('username')
            um.Ornganization_id_id = request.POST.get('organization')
            um.Mobile_no = request.POST.get('phone')
            um.Designation = request.POST.get('designation')
            um.Gender = request.POST.get('gender')
            um.Age = request.POST.get('age')
            um.Role = "SU"
            um.Status = "Active"
            # check = userMaster.objects.filter()
            um.save()
            lg = loginMaster()
            lg.Username = request.POST.get('username')
            lg.Password = request.POST.get('phone')
            lg.Role = "SU"
            lg.save()
            # data=usetrmaster.objects.get(username=request.session['usid'])
            # um.Ornganization_id_id=data.organization_id
            messages.success(request, "SuperUser added successfully.")
            return redirect('/add_superUser')
def delete_superUser(request):
    if 'uid' not in request.session:
        return redirect('/login')
    else:
        rem = userMaster.objects.all()
        return render(request,'Delete_superuser.html',{'rem':rem})
def delete_superUser_db(request,User_id):
    if 'uid' not in request.session:
        return redirect('/login')
    else:
        delete1 = userMaster.objects.get(User_id=User_id)
        lg = loginMaster.objects.get(Username = delete1.Username,Password = delete1.Mobile_no)
        delete1.delete()
        lg.delete()
        messages.success(request,"SuperUser removed successfully")
        return redirect('/delete_superUser')
def superUser_home(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        current_date = timezone.now()
        dt = userMaster.objects.get(User_id = request.session['usid'])
        orgid = dt.Ornganization_id_id
        this_month_start = current_date.replace(day=1,hour=0,minute=0,second=0,microsecond=0)
        this_month_donation = donationDetails.objects.filter(Donation_date__range = [this_month_start,current_date]).filter(Ornganization_id_id=orgid).filter(Donation_status = 'Donated').count()
        
        #lastmonth donation
        last_month_start = (this_month_start - timedelta(days=1)).replace(day=1)
        last_month_end = this_month_start - timedelta(days=1)
        last_month_donations = donationDetails.objects.filter(Donation_date__range = [last_month_start,last_month_end]).filter(Ornganization_id_id=orgid).filter(Donation_status='Donated').count()
        
        #this year
        this_year_start = current_date.replace(month=1,day=1,hour=0,minute=0,second=0,microsecond=0)
        this_year_donations = donationDetails.objects.filter(Donation_date__range = [this_year_start,current_date]).filter(Ornganization_id_id=orgid).filter(Donation_status = 'Donated').count()
        
        
        return render(request,'Superuser_home.html',{'thismonth':this_month_donation,'lastmonth':last_month_donations,'thisyear':this_year_donations})
def login(request):
    return render(request,'login.html')
def login_db(request):
    data = loginMaster.objects.all()
    username = request.POST.get('username')
    password = request.POST.get('password')
    flag = 0
    for da in data:
        if username == da.Username and password == da.Password:
            type1 = da.Role
            flag = 1
            if type1 == 'admin':
                request.session['uid'] = username
                return redirect('/admin_home')
            elif type1 == 'SU':
                sup = userMaster.objects.get(Username = username)
                request.session['usid'] = sup.User_id
                return redirect('/superUser_home') 
            elif type1 == 'Privilaged User':
                pu = userMaster.objects.get(Username = username)
                request.session['puid'] = pu.User_id
                return redirect('/PU_home')
            elif type1 == 'User':
                us = userMaster.objects.get(Username = username)
                request.session['user'] = us.User_id
                return redirect('/USER_home')
            else:
                messages.error(request,"Invalid")
                return render(request,'login.html',{'error': "Invalid"})
    if flag == 0:
        messages.success(request,"Invalid Username or Password")
        return render(request,'login.html', {'error': "Invalid username or password"})
def lgout(request):
    del request.session['uid']
    return redirect('/login')
def lgout_Superuser(request):
    del request.session['usid']
    return redirect('/login')
def lgout_PU(request):
    del request.session['puid']
    return redirect('/login')
def lgout_USER(request):
    del request.session['user']
    return redirect('/login')
def add_user(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        return render(request,'Add_user.html')
def add_user_db(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            usm = userMaster()
            usm.Username = request.POST.get('username')
            dat = userMaster.objects.get(User_id = request.session['usid'])
            usm.Ornganization_id_id = dat.Ornganization_id_id
            usm.Mobile_no = request.POST.get('phone')
            usm.Designation = request.POST.get('designation')
            usm.Gender = request.POST.get('gender')
            usm.Age = request.POST.get('age')
            usm.Role = request.POST.get('role')
            usm.Status = "Active"
            usm.save()
            lgm = loginMaster()
            lgm.Username = request.POST.get('username')
            lgm.Password = request.POST.get('phone')
            lgm.Role = request.POST.get('role')
            lgm.save()
            messages.success(request, "User added successfully.")
            return redirect('/add_user')
def delete_user(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        dat = userMaster.objects.get(User_id = request.session['usid'])
        orid = dat.Ornganization_id_id
        # print('dat',dat)
        # print('orid',orid)
        datas = userMaster.objects.filter(Ornganization_id_id = orid).exclude(Role='SU')
        # print('datas',datas)
        return render(request,'delete_user.html',{'datas':datas})
def delete_user_db(request,User_id):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        dt = userMaster.objects.get(User_id=User_id)
        a = dt.Role
        lg1 = loginMaster.objects.get(Username = dt.Username,Password = dt.Mobile_no)
        dt.delete()
        lg1.delete() 
        messages.success(request,f"{a} removed successfully")
        return redirect('/delete_user')
def add_area(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        return render(request,'Add_area.html')
def add_area_db(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            ar = AreaMaster()
            ar.Area_name = request.POST.get('area')
            ar.Status = "Yes"
            dt = userMaster.objects.get(User_id = request.session['usid'])
            ar.Ornganization_id_id = dt.Ornganization_id_id
            ar.save()
            messages.success(request, "Area added successfully.")
            return redirect('/add_area')
def delete_area(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        dat = userMaster.objects.get(User_id = request.session['usid'])
        orid = dat.Ornganization_id_id
        datas = AreaMaster.objects.filter(Ornganization_id_id = orid)
        return render(request,'delete_area.html',{'datas':datas})
def delete_area_db(request,Area_code):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        dl = AreaMaster.objects.get(Area_code=Area_code)
        a = dl.Area_name
        dl.delete()
        messages.success(request,f" Area {a} removed successfully")
        return redirect('/delete_area')
def add_hospital_superUser(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        return render(request,'add_hospital_superUser.html')
def add_hospital_superUser_db(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            hos = hospitalMaster()
            hos.Hospital_name = request.POST.get('hospital')
            hos.Address = request.POST.get('address')
            hos.Status = "Yes"
            dt = userMaster.objects.get(User_id = request.session['usid'])
            hos.Ornganization_id_id = dt.Ornganization_id_id
            hos.save()
            messages.success(request, "Hospital added successfully.")
            return redirect('/add_hospital_superUser')
def delete_hospital_superUser(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        dt = userMaster.objects.get(User_id = request.session['usid'])
        orid = dt.Ornganization_id_id
        datas = hospitalMaster.objects.filter(Ornganization_id_id = orid)
        return render(request,'delete_hospital_superUser.html',{'datas':datas})
def delete_hospital_superUser_db(request,Hospital_id):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        dle = hospitalMaster.objects.get(Hospital_id = Hospital_id)
        a = dle.Hospital_name
        dle.delete()
        messages.success(request,f" {a} Hospital removed successfully")
        return redirect('/delete_hospital_superUser')
def Donor_registration_SU(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        dt = userMaster.objects.get(User_id = request.session['usid'])
        orgid = dt.Ornganization_id_id
        datas = AreaMaster.objects.filter(Ornganization_id_id = orgid)
        return render(request,'Donor_registration_SU.html',{'datas':datas})
def Donor_registration_SU_db(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            dnm = donorMaster()
            dnm.Donor_name = request.POST.get('donor')
            dnm.Gender = request.POST.get('gender')
            dnm.DOB = request.POST.get('dob')
            dnm.Age = request.POST.get('age')
            dnm.Address = request.POST.get('address')
            # dnm.Area = request.POST.get('area')
            area_id = request.POST.get('area')
            dnm.Area = AreaMaster.objects.get(Area_code = area_id)
            dnm.Phone_no = request.POST.get('phone')
            dnm.Blood_group = request.POST.get('bloodgroup')
            dnm.Donated_earlier = request.POST.get('donated') == 'yes'
            dnm.Last_donation_date = request.POST.get('lastdate') if dnm.Donated_earlier else None
            dnm.Created_date = date.today().strftime("%Y-%m-%d")
            dnm.Status = "Yes"
            dt = userMaster.objects.get(User_id = request.session['usid'])
            dnm.Ornganization_id_id = dt.Ornganization_id_id
            dnm.Captured_by = dt
            action = request.POST.get("action")
            if action == 'save':
                dnm.save()
                messages.success(request, "Donor registered successfully.")
                return redirect('/Donor_registration_SU')
            elif action == 'save_donate':
                dnm.save()
                don = dnm
                dat1 = userMaster.objects.get(User_id = request.session['usid'])
                orid = dat1.Ornganization_id_id
                datas = hospitalMaster.objects.filter(Ornganization_id_id = orid)
                return render(request,'Donate_blood_SU.html',{'donor':don,'datas': datas})
def donate_blood_db(request,Donor_id):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            dd = donationDetails()
            dt1 = userMaster.objects.get(User_id = request.session['usid'])
            dd.Ornganization_id_id = dt1.Ornganization_id_id
            dd.Captured_by = dt1
            hos_id = request.POST.get('hospital')
            dd.Hospital_id = hospitalMaster.objects.get(Hospital_id = hos_id)
            a = donorMaster.objects.get(Donor_id = Donor_id)
            dd.Donor_id = a
            dd.Patient_bystander = request.POST.get('patientName')
            dd.Contact_no = request.POST.get('contactNo')
            dd.Purpose = request.POST.get('purpose')
            dd.Donation_date = request.POST.get('donationsdate')
            dd.Donation_status = request.POST.get('donationStatus')
            dd.Created_date = date.today().strftime("%Y-%m-%d")
            dd.Status = "Yes"
            dd.save()
            if dd.Donation_status == 'Donated':
                a.Last_donation_date = dd.Donation_date
                a.Status = 'No'
                a.save()
            messages.success(request,"Save and donated blood successfully")
            return redirect('/Donor_registration_SU')
def donor_search_SU(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        today = date.today()
        
        dt = userMaster.objects.get(User_id = request.session['usid'])
        orgid = dt.Ornganization_id_id
        ineligible_donors = donorMaster.objects.filter(Status = 'No',Ornganization_id_id=orgid)
        for donor in ineligible_donors:
            try:
                last_donation = datetime.datetime.strptime(donor.Last_donation_date, "%Y-%m-%d").date()
                days_since_donation = (today - last_donation).days
                if days_since_donation >= 90:
                    donor.Status = 'Yes'
                    donor.save()
            except Exception as e:
                print(f"Error updating donor status:{str(e)}")
                pass
        datas = AreaMaster.objects.filter(Ornganization_id_id = orgid)
        return render(request,'Donor_search_SU.html',{'datas':datas})
def donor_search_SU_db(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            dt = userMaster.objects.get(User_id = request.session['usid'])
            orgid = dt.Ornganization_id_id
            datas = AreaMaster.objects.filter(Ornganization_id_id = orgid)
            today = date.today()
            ineligible_donors = donorMaster.objects.filter(Status = 'No',Ornganization_id_id = orgid)
            for donor in ineligible_donors:
                try:
                    last_donation = datetime.datetime.strptime(donor.Last_donation_date,"%Y-%m-%d").date()
                    day_since_donation = (today-last_donation).days
                    if day_since_donation >= 90:
                        donor.Status = 'Yes'
                        donor.save()
                except Exception as e:
                    print(f"Error updating donor status: {str(e)}")
                    pass
            if request.POST.get('donor_name') != '':
                a = request.POST.get('donor_name')
                data = donorMaster.objects.filter(Donor_name = str(a)).filter(Ornganization_id_id = orgid).filter(Status = "Yes")
                return render(request,'Donor_search_SU.html',{'donors':data,'datas':datas})
            if request.POST.get('mobile') != '':
                b = request.POST.get('mobile')
                data = donorMaster.objects.filter(Phone_no = b).filter(Ornganization_id_id = orgid).filter(Status = "Yes")
                return render(request,'Donor_search_SU.html',{'donors':data,'datas':datas})
            if request.POST.get('blood_group') and request.POST.get('area') != '':
                c = request.POST.get('blood_group')
                d = request.POST.get('area')
                data = donorMaster.objects.filter(Blood_group = c).filter(Area_id = d).filter(Ornganization_id_id = orgid).filter(Status = "Yes")
                return render(request,'Donor_search_SU.html',{'donors':data,'datas':datas})
            if request.POST.get('blood_group') != '':
                c = request.POST.get('blood_group')
                data = donorMaster.objects.filter(Blood_group = c).filter(Ornganization_id_id = orgid).filter(Status = "Yes")
                return render(request,'Donor_search_SU.html',{'donors':data,'datas':datas})
            if request.POST.get('area') != '':
                d = request.POST.get('area')
                data = donorMaster.objects.filter(Area_id = d).filter(Ornganization_id_id = orgid).filter(Status = "Yes")
                return render(request,'Donor_search_SU.html',{'donors':data,'datas':datas})
            
        # messages.error(request,"No data found")
        return render(request, 'Donor_search_SU.html', {'datas': datas})
def donate_blood_SU(request,Donor_id):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        dt = userMaster.objects.get(User_id = request.session['usid'])
        orid = dt.Ornganization_id_id
        datas1 = donorMaster.objects.get(Donor_id = Donor_id,Ornganization_id_id = orid)
        # print(datas1)
        datas = hospitalMaster.objects.filter(Ornganization_id_id = orid)
        return render(request,'Donate_blood_SU2.html',{'donor':datas1,'datas':datas})
def donate_blood_SU_db(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            dd = donationDetails()
            dt1 = userMaster.objects.get(User_id = request.session['usid'])
            dd.Ornganization_id_id = dt1.Ornganization_id_id
            dd.Captured_by = dt1
            hos_id = request.POST.get('hospital')
            dd.Hospital_id = hospitalMaster.objects.get(Hospital_id = hos_id)
            a = request.POST.get('donorId')
            b = donorMaster.objects.get(Donor_id = a)
            dd.Donor_id = b
            dd.Patient_bystander = request.POST.get('patientName')
            dd.Contact_no = request.POST.get('contactNo')
            dd.Purpose = request.POST.get('purpose')
            dd.Donation_date = request.POST.get('donationsdate')
            dd.Donation_status = request.POST.get('donationStatus')
            dd.Created_date = date.today().strftime("%Y-%m-%d")
            dd.Status = "Yes"
            dd.save()
            if dd.Donation_status == 'Donated':
                b.Last_donation_date = dd.Donation_date
                b.Status = 'No'
                b.save()
            messages.success(request,"Blood donated successfully")
            return redirect('/donor_search_SU')
        
def Donor_edit_SU(request,Donor_id):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        dt = userMaster.objects.get(User_id = request.session['usid'])
        orid = dt.Ornganization_id_id
        datas1 = donorMaster.objects.get(Donor_id = Donor_id,Ornganization_id_id = orid)
        datas = AreaMaster.objects.filter(Ornganization_id_id=orid)
        return render(request,'Donor_edit_SU.html',{'datas1':datas1,'datas':datas}) 
def Donor_edit_SU_db(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            donor_id = request.POST.get('donor_id')
            dt = userMaster.objects.get(User_id = request.session['usid'])
            orid = dt.Ornganization_id_id
            dd = donorMaster.objects.get(Donor_id=donor_id)
            dd.Donor_name = request.POST.get('donor')
            dd.Phone_no = request.POST.get('phone')
            dd.Address = request.POST.get('address')
            area_id = request.POST.get('area')
            dd.Area = AreaMaster.objects.get(Area_code = area_id) 
            dd.save()
            messages.success(request,"Donor details edited successfully")
            return redirect('/donor_search_SU')
def view_donors_SU(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        dt = userMaster.objects.get(User_id = request.session['usid'])
        orgid = dt.Ornganization_id_id
        datas = AreaMaster.objects.filter(Ornganization_id_id = orgid)
        return render(request,'Donor_view_SU.html',{'datas':datas})
def view_donors_db_SU(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':  
            dt = userMaster.objects.get(User_id = request.session['usid'])
            orgid = dt.Ornganization_id_id
            datas = AreaMaster.objects.filter(Ornganization_id_id = orgid)
            if request.POST.get('donor_name') != '':
                a = request.POST.get('donor_name')
                data = donorMaster.objects.filter(Donor_name = str(a)).filter(Ornganization_id_id = orgid)
                return render(request,'Donor_view_SU.html',{'donors':data,'datas':datas})  
            if request.POST.get('mobile') != '':
                b = request.POST.get('mobile') 
                data = donorMaster.objects.filter(Phone_no = b).filter(Ornganization_id_id = orgid)
                return render(request,'Donor_view_SU.html',{'donors':data,'datas':datas})
            if request.POST.get('blood_group') != '' and request.POST.get('area') != '':
                f = request.POST.get('blood_group')
                g = request.POST.get('area')
                data = donorMaster.objects.filter(Blood_group = f).filter(Area_id = int(g)).filter(Ornganization_id_id = orgid)
                return render(request,'Donor_view_SU.html',{'donors':data,'datas':datas})
            if request.POST.get('blood_group') != '':
                c = request.POST.get('blood_group')
                data = donorMaster.objects.filter(Blood_group = c).filter(Ornganization_id_id = orgid)
                return render(request,'Donor_view_SU.html',{'donors':data,'datas':datas})
            if request.POST.get('area') != '':
                d = request.POST.get('area')
                data = donorMaster.objects.filter(Area_id = d).filter(Ornganization_id_id = orgid)
                return render(request,'Donor_view_SU.html',{'donors':data,'datas':datas})
              
        return render(request, 'Donor_view_SU.html', {'datas': datas}) 
def Donor_report_SU(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        return render(request,'Donor_report_SU.html')
def Donor_report_find(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            fromdate = request.POST.get('fromdate')
            todate = request.POST.get('todate')
            
            request.session['fromdate'] = fromdate
            request.session['todate'] = todate
            
            
            dt = userMaster.objects.get(User_id = request.session['usid'])
            orgid = dt.Ornganization_id_id
            donation_details = donationDetails.objects.filter(Donation_date__range=[fromdate, todate]).filter(Ornganization_id_id = orgid).filter(Donation_status = 'Donated')
            return render(request, 'Donor_report_SU.html',{'donors':donation_details})
        return redirect('/Donor_report_SU')
def export_donor_report_SU(request):
    if 'usid' not in request.session:
        return redirect('/login')
    else:
        fromdate = request.session.get('fromdate')
        todate = request.session.get('todate')
        
        dt = userMaster.objects.get(User_id=request.session['usid'])
        orgid = dt.Ornganization_id_id
        donation_details = donationDetails.objects.filter(Donation_date__range=[fromdate, todate]).filter(Ornganization_id_id=orgid).filter(Donation_status='Donated')
        wb = Workbook()
        ws = wb.active
        ws.title = "Donation Report"
        headers = ['Donor Name', 'Donation Date', 'Hospital', 'Blood Group', 'Patient']
        for col,header in enumerate(headers,1):
            cell = ws.cell(row=1,column=col,value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row,donation in enumerate(donation_details,2):
            ws.cell(row=row,column=1,value=donation.Donor_id.Donor_name)
            ws.cell(row=row, column=2, value=str(donation.Donation_date))
            ws.cell(row=row, column=3, value=donation.Hospital_id.Hospital_name)
            ws.cell(row=row, column=4, value=donation.Donor_id.Blood_group)
            ws.cell(row=row, column=5, value=donation.Patient_bystander)
        
        for col in range(1,len(headers)+1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].auto_size = True
            
        response = HttpResponse(content_type='application/vandapplication/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Donor_Report.xlsx'
        wb.save(response)
        return response                                     
def PU_home(request):
    if 'puid' not in request.session:
        return redirect('/login')
    else:
        current_date = timezone.now()
        dt = userMaster.objects.get(User_id = request.session['puid'])
        orgid = dt.Ornganization_id_id
        this_month_start = current_date.replace(day=1,hour=0,minute=0,second=0,microsecond=0)
        this_month_donation = donationDetails.objects.filter(Donation_date__range = [this_month_start,current_date]).filter(Ornganization_id_id=orgid).filter(Donation_status = 'Donated').count()
        
        #lastmonth donation
        last_month_start = (this_month_start - timedelta(days=1)).replace(day=1)
        last_month_end = this_month_start - timedelta(days=1)
        last_month_donations = donationDetails.objects.filter(Donation_date__range = [last_month_start,last_month_end]).filter(Ornganization_id_id=orgid).filter(Donation_status='Donated').count()
        
        #this year
        this_year_start = current_date.replace(month=1,day=1,hour=0,minute=0,second=0,microsecond=0)
        this_year_donations = donationDetails.objects.filter(Donation_date__range = [this_year_start,current_date]).filter(Ornganization_id_id=orgid).filter(Donation_status = 'Donated').count()
        return render(request,'PU_home.html',{'thismonth':this_month_donation,'lastmonth':last_month_donations,'thisyear':this_year_donations})
def PU_add_hospital(request):
    if 'puid' not in request.session:
        return redirect('/login')
    else:
        return render(request,'PU_add_hospital.html')
def PU_add_hospital_db(request):
    if 'puid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            hos = hospitalMaster()
            hos.Hospital_name = request.POST.get('hospital')
            hos.Address = request.POST.get('address')
            hos.Status = "Yes"
            dt = userMaster.objects.get(User_id = request.session['puid'])
            hos.Ornganization_id_id = dt.Ornganization_id_id
            hos.save()
            messages.success(request, "Hospital added successfully.")
            return redirect('/PU_add_hospital')
def PU_view_hospital(request):
    if 'puid' not in request.session:
        return redirect('/login')
    else:
        dt = userMaster.objects.get(User_id = request.session['puid'])
        orid = dt.Ornganization_id_id
        datas = hospitalMaster.objects.filter(Ornganization_id_id = orid)
        return render(request,'PU_view_hospital.html',{'datas':datas})
def PU_donor_registration(request):
    if 'puid' not in request.session:
        return redirect('/login')
    else:
        dt = userMaster.objects.get(User_id = request.session['puid'])
        orgid = dt.Ornganization_id_id
        datas = AreaMaster.objects.filter(Ornganization_id_id = orgid)
        return render(request,'PU_donor_registration.html',{'datas':datas})
def PU_donor_registration_db(request):
    if 'puid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            dnm = donorMaster()
            dnm.Donor_name = request.POST.get('donor')
            dnm.Gender = request.POST.get('gender')
            dnm.DOB = request.POST.get('dob')
            dnm.Age = request.POST.get('age')
            dnm.Address = request.POST.get('address')
            # dnm.Area = request.POST.get('area')
            area_id = request.POST.get('area')
            dnm.Area = AreaMaster.objects.get(Area_code = area_id)
            dnm.Phone_no = request.POST.get('phone')
            dnm.Blood_group = request.POST.get('bloodgroup')
            dnm.Donated_earlier = request.POST.get('donated') == 'yes'
            dnm.Last_donation_date = request.POST.get('lastdate') if dnm.Donated_earlier else None
            dnm.Created_date = date.today().strftime("%Y-%m-%d")
            dnm.Status = "Yes"
            dt = userMaster.objects.get(User_id = request.session['puid'])
            dnm.Ornganization_id_id = dt.Ornganization_id_id
            dnm.Captured_by = dt
            action = request.POST.get("action")
            if action == 'save':
                dnm.save()
                messages.success(request, "Donor registered successfully.")
                return redirect('/PU_donor_registration')
            elif action == 'save_donate':
                dnm.save()
                don = dnm
                dat1 = userMaster.objects.get(User_id = request.session['puid'])
                orid = dat1.Ornganization_id_id
                datas = hospitalMaster.objects.filter(Ornganization_id_id = orid)
                return render(request,'PU_donate_blood.html',{'donor':don,'datas': datas})  
def PU_donate_blood_db(request,Donor_id):
    if 'puid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            dd = donationDetails()
            dt1 = userMaster.objects.get(User_id = request.session['puid'])
            dd.Ornganization_id_id = dt1.Ornganization_id_id
            dd.Captured_by = dt1
            hos_id = request.POST.get('hospital')
            dd.Hospital_id = hospitalMaster.objects.get(Hospital_id = hos_id)
            a = donorMaster.objects.get(Donor_id = Donor_id)
            dd.Donor_id = a
            dd.Patient_bystander = request.POST.get('patientName')
            dd.Contact_no = request.POST.get('contactNo')
            dd.Purpose = request.POST.get('purpose')
            dd.Donation_date = request.POST.get('donationsdate')
            dd.Donation_status = request.POST.get('donationStatus')
            dd.Created_date = date.today().strftime("%Y-%m-%d")
            dd.Status = "Yes"
            dd.save()
            if dd.Donation_status == 'Donated':
                a.Last_donation_date = dd.Donation_date
                a.Status = 'No'
                a.save()
            messages.success(request,"Save and donated blood successfully")
            return redirect('/PU_donor_registration')
def PU_donor_search(request):
    if 'puid' not in request.session:
        return redirect('/login')
    else:
        today = date.today()
        
        dt = userMaster.objects.get(User_id = request.session['puid'])
        orgid = dt.Ornganization_id_id
        ineligible_donors = donorMaster.objects.filter(Status = 'No',Ornganization_id_id=orgid)
        for donor in ineligible_donors:
            try:
                last_donation = datetime.datetime.strptime(donor.Last_donation_date, "%Y-%m-%d").date()
                days_since_donation = (today - last_donation).days
                if days_since_donation >= 90:
                    donor.Status = 'Yes'
                    donor.save()
            except Exception as e:
                print(f"Error updating donor status:{str(e)}")
                pass
        datas = AreaMaster.objects.filter(Ornganization_id_id = orgid)
        return render(request,'PU_donor_search.html',{'datas':datas})
def PU_donor_search_db(request):
    if 'puid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            dt = userMaster.objects.get(User_id = request.session['puid'])
            orgid = dt.Ornganization_id_id
            datas = AreaMaster.objects.filter(Ornganization_id_id = orgid)
            today = date.today()
            ineligible_donors = donorMaster.objects.filter(Status = 'No',Ornganization_id_id = orgid)
            for donor in ineligible_donors:
                try:
                    last_donation = datetime.datetime.strptime(donor.Last_donation_date,"%Y-%m-%d").date()
                    day_since_donation = (today-last_donation).days
                    if day_since_donation >= 90:
                        donor.Status = 'Yes'
                        donor.save()
                except Exception as e:
                    print(f"Error updating donor status: {str(e)}")
                    pass
            if request.POST.get('donor_name') != '':
                a = request.POST.get('donor_name')
                data = donorMaster.objects.filter(Donor_name = str(a)).filter(Ornganization_id_id = orgid).filter(Status = 'Yes')
                return render(request,'PU_donor_search.html',{'donors':data,'datas':datas})  
            if request.POST.get('mobile') != '':
                b = request.POST.get('mobile') 
                data = donorMaster.objects.filter(Phone_no = b).filter(Ornganization_id_id = orgid).filter(Status = 'Yes')
                return render(request,'PU_donor_search.html',{'donors':data,'datas':datas})
            if request.POST.get('blood_group') != '' and request.POST.get('area') != '':
                f = request.POST.get('blood_group')
                g = request.POST.get('area')
                data = donorMaster.objects.filter(Blood_group = f).filter(Area_id = int(g)).filter(Ornganization_id_id = orgid).filter(Status = 'Yes')
                return render(request,'PU_donor_search.html',{'donors':data,'datas':datas})
            if request.POST.get('blood_group') != '':
                c = request.POST.get('blood_group')
                data = donorMaster.objects.filter(Blood_group = c).filter(Ornganization_id_id = orgid).filter(Status = 'Yes')
                return render(request,'PU_donor_search.html',{'donors':data,'datas':datas})
            if request.POST.get('area') != '':
                d = request.POST.get('area')
                data = donorMaster.objects.filter(Area_id = d).filter(Ornganization_id_id = orgid).filter(Status = 'Yes')
                return render(request,'PU_donor_search.html',{'donors':data,'datas':datas})
              
        return render(request, 'PU_donor_search.html', {'datas': datas})    
def PU2_donate_blood(request,Donor_id):
    if 'puid' not in request.session:
        return redirect('/login')
    else:
        dt = userMaster.objects.get(User_id = request.session['puid'])
        orid = dt.Ornganization_id_id
        datas1 = donorMaster.objects.get(Donor_id = Donor_id,Ornganization_id_id = orid)
        # print(datas1)
        datas = hospitalMaster.objects.filter(Ornganization_id_id = orid)
        return render(request,'PU2_donate_blood.html',{'donor':datas1,'datas':datas}) 
def PU2_donate_blood_db(request):
    if 'puid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            
            dd = donationDetails()
            dt1 = userMaster.objects.get(User_id = request.session['puid'])
            dd.Ornganization_id_id = dt1.Ornganization_id_id
            dd.Captured_by = dt1
            hos_id = request.POST.get('hospital')
            dd.Hospital_id = hospitalMaster.objects.get(Hospital_id = hos_id)
            a = request.POST.get('donorId')
            b = donorMaster.objects.get(Donor_id = a)
            dd.Donor_id = b
            dd.Patient_bystander = request.POST.get('patientName')
            dd.Contact_no = request.POST.get('contactNo')
            dd.Purpose = request.POST.get('purpose')
            dd.Donation_date = request.POST.get('donationsdate')
            dd.Donation_status = request.POST.get('donationStatus')
            dd.Created_date = date.today().strftime("%Y-%m-%d")
            dd.Status = "Yes"
            dd.save()
            if dd.Donation_status == 'Donated':
                b.Last_donation_date = dd.Donation_date
                b.Status = 'No'
                b.save()
            messages.success(request,"Blood donated successfully")
            return redirect('/PU_donor_search')
def PU_donor_edit(request,Donor_id):
    if 'puid' not in request.session:
        return redirect('/login')
    else:
        dt = userMaster.objects.get(User_id = request.session['puid'])
        orid = dt.Ornganization_id_id
        datas1 = donorMaster.objects.get(Donor_id = Donor_id,Ornganization_id_id = orid)
        datas = AreaMaster.objects.filter(Ornganization_id_id=orid)
        return render(request,'PU_donor_edit.html',{'datas1':datas1,'datas':datas})
def PU_donor_edit_db(request):
    if 'puid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            donor_id = request.POST.get('donor_id')
            dt = userMaster.objects.get(User_id = request.session['puid'])
            orid = dt.Ornganization_id_id
            dd = donorMaster.objects.get(Donor_id=donor_id)
            dd.Donor_name = request.POST.get('donor')
            dd.Phone_no = request.POST.get('phone')
            dd.Address = request.POST.get('address')
            area_id = request.POST.get('area')
            dd.Area = AreaMaster.objects.get(Area_code = area_id) 
            dd.save()
            messages.success(request,"Donor details edited successfully")
            return redirect('/PU_donor_search')
def PU_view_donors(request):
    if 'puid' not in request.session:
        return redirect('/login')
    else:
        dt = userMaster.objects.get(User_id = request.session['puid'])
        orgid = dt.Ornganization_id_id
        datas = AreaMaster.objects.filter(Ornganization_id_id = orgid)
        return render(request,'PU_view_donors.html',{'datas':datas})
def PU_view_donors_db(request):
    if 'puid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':  
            dt = userMaster.objects.get(User_id = request.session['puid'])
            orgid = dt.Ornganization_id_id
            datas = AreaMaster.objects.filter(Ornganization_id_id = orgid)
            if request.POST.get('donor_name') != '':
                a = request.POST.get('donor_name')
                data = donorMaster.objects.filter(Donor_name = str(a)).filter(Ornganization_id_id = orgid)
                return render(request,'PU_view_donors.html',{'donors':data,'datas':datas})  
            if request.POST.get('mobile') != '':
                b = request.POST.get('mobile') 
                data = donorMaster.objects.filter(Phone_no = b).filter(Ornganization_id_id = orgid)
                return render(request,'PU_view_donors.html',{'donors':data,'datas':datas})
            if request.POST.get('blood_group') != '' and request.POST.get('area') != '':
                f = request.POST.get('blood_group')
                g = request.POST.get('area')
                data = donorMaster.objects.filter(Blood_group = f).filter(Area_id = int(g)).filter(Ornganization_id_id = orgid)
                return render(request,'PU_view_donors.html',{'donors':data,'datas':datas})
            if request.POST.get('blood_group') != '':
                c = request.POST.get('blood_group')
                data = donorMaster.objects.filter(Blood_group = c).filter(Ornganization_id_id = orgid)
                return render(request,'PU_view_donors.html',{'donors':data,'datas':datas})
            if request.POST.get('area') != '':
                d = request.POST.get('area')
                data = donorMaster.objects.filter(Area_id = d).filter(Ornganization_id_id = orgid)
                return render(request,'PU_view_donors.html',{'donors':data,'datas':datas})
              
        return render(request, 'PU_view_donors.html', {'datas': datas}) 
def PU_donor_report(request):
    if 'puid' not in request.session:
        return redirect('/login')
    else:
        return render(request,'PU_donor_report.html')
def PU_donor_report_find(request):
    if 'puid' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            fromdate = request.POST.get('fromdate')
            todate = request.POST.get('todate')
            
            request.session['fromdate'] = fromdate
            request.session['todate'] = todate
            
            
            dt = userMaster.objects.get(User_id = request.session['puid'])
            orgid = dt.Ornganization_id_id
            donation_details = donationDetails.objects.filter(Donation_date__range=[fromdate, todate]).filter(Ornganization_id_id = orgid).filter(Donation_status = 'Donated')
            return render(request, 'PU_donor_report.html',{'donors':donation_details})
        return redirect('/PU_donor_report') 
def PU_export_donor_report(request):
    if 'puid' not in request.session:
        return redirect('/login')
    else:
        fromdate = request.session.get('fromdate')
        todate = request.session.get('todate')
        
        dt = userMaster.objects.get(User_id=request.session['puid'])
        orgid = dt.Ornganization_id_id
        donation_details = donationDetails.objects.filter(Donation_date__range=[fromdate, todate]).filter(Ornganization_id_id=orgid).filter(Donation_status='Donated')
        wb = Workbook()
        ws = wb.active
        ws.title = "Donation Report"
        headers = ['Donor Name', 'Donation Date', 'Hospital', 'Blood Group', 'Patient']
        for col,header in enumerate(headers,1):
            cell = ws.cell(row=1,column=col,value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row,donation in enumerate(donation_details,2):
            ws.cell(row=row,column=1,value=donation.Donor_id.Donor_name)
            ws.cell(row=row, column=2, value=str(donation.Donation_date))
            ws.cell(row=row, column=3, value=donation.Hospital_id.Hospital_name)
            ws.cell(row=row, column=4, value=donation.Donor_id.Blood_group)
            ws.cell(row=row, column=5, value=donation.Patient_bystander)
        
        for col in range(1,len(headers)+1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].auto_size = True
            
        response = HttpResponse(content_type='application/vandapplication/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Donor_Report.xlsx'
        wb.save(response)
        return response         
def USER_home(request):
    if 'user' not in request.session:
        return redirect('/login')
    else:
        current_date = timezone.now()
        dt = userMaster.objects.get(User_id = request.session['user'])
        orgid = dt.Ornganization_id_id
        this_month_start = current_date.replace(day=1,hour=0,minute=0,second=0,microsecond=0)
        this_month_donation = donationDetails.objects.filter(Donation_date__range = [this_month_start,current_date]).filter(Ornganization_id_id=orgid).filter(Donation_status = 'Donated').count()
        
        #lastmonth donation
        last_month_start = (this_month_start - timedelta(days=1)).replace(day=1)
        last_month_end = this_month_start - timedelta(days=1)
        last_month_donations = donationDetails.objects.filter(Donation_date__range = [last_month_start,last_month_end]).filter(Ornganization_id_id=orgid).filter(Donation_status='Donated').count()
        
        #this year
        this_year_start = current_date.replace(month=1,day=1,hour=0,minute=0,second=0,microsecond=0)
        this_year_donations = donationDetails.objects.filter(Donation_date__range = [this_year_start,current_date]).filter(Ornganization_id_id=orgid).filter(Donation_status = 'Donated').count()
        return render(request,'USER_home.html',{'thismonth':this_month_donation,'lastmonth':last_month_donations,'thisyear':this_year_donations})
def USER_donor_registration(request):
    if 'user' not in request.session:
        return redirect('/login')
    else:
        dt = userMaster.objects.get(User_id = request.session['user'])
        orgid = dt.Ornganization_id_id
        datas = AreaMaster.objects.filter(Ornganization_id_id = orgid)
        return render(request,'USER_donor_registration.html',{'datas':datas})
def USER_donor_registration_db(request):
    if 'user' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            dnm = donorMaster()
            dnm.Donor_name = request.POST.get('donor')
            dnm.Gender = request.POST.get('gender')
            dnm.DOB = request.POST.get('dob')
            dnm.Age = request.POST.get('age')
            dnm.Address = request.POST.get('address')
            # dnm.Area = request.POST.get('area')
            area_id = request.POST.get('area')
            dnm.Area = AreaMaster.objects.get(Area_code = area_id)
            dnm.Phone_no = request.POST.get('phone')
            dnm.Blood_group = request.POST.get('bloodgroup')
            dnm.Donated_earlier = request.POST.get('donated') == 'yes'
            dnm.Last_donation_date = request.POST.get('lastdate') if dnm.Donated_earlier else None
            dnm.Created_date = date.today().strftime("%Y-%m-%d")
            dnm.Status = "Yes"
            dt = userMaster.objects.get(User_id = request.session['user'])
            dnm.Ornganization_id_id = dt.Ornganization_id_id
            dnm.Captured_by = dt
            action = request.POST.get("action")
            if action == 'save':
                dnm.save()
                messages.success(request, "Donor registered successfully.")
                return redirect('/USER_donor_registration')
            elif action == 'save_donate':
                dnm.save()
                don = dnm
                dat1 = userMaster.objects.get(User_id = request.session['user'])
                orid = dat1.Ornganization_id_id
                datas = hospitalMaster.objects.filter(Ornganization_id_id = orid)
                return render(request,'USER_donate_blood.html',{'donor':don,'datas': datas})
def USER_donate_blood_db(request,Donor_id):
    if 'user' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            dd = donationDetails()
            dt1 = userMaster.objects.get(User_id = request.session['user'])
            dd.Ornganization_id_id = dt1.Ornganization_id_id
            dd.Captured_by = dt1
            hos_id = request.POST.get('hospital')
            dd.Hospital_id = hospitalMaster.objects.get(Hospital_id = hos_id)
            a = donorMaster.objects.get(Donor_id = Donor_id)
            dd.Donor_id = a
            dd.Patient_bystander = request.POST.get('patientName')
            dd.Contact_no = request.POST.get('contactNo')
            dd.Purpose = request.POST.get('purpose')
            dd.Donation_date = request.POST.get('donationsdate')
            dd.Donation_status = request.POST.get('donationStatus')
            dd.Created_date = date.today().strftime("%Y-%m-%d")
            dd.Status = "Yes"
            dd.save()
            if dd.Donation_status == 'Donated':
                a.Last_donation_date = dd.Donation_date
                a.Status = 'No'
                a.save()
            messages.success(request,"Save and donated blood successfully")
            return redirect('/USER_donor_registration')
def USER_donor_search(request):
    if 'user' not in request.session:
        return redirect('/login')
    else:
        today = date.today()
        
        dt = userMaster.objects.get(User_id = request.session['user'])
        orgid = dt.Ornganization_id_id
        ineligible_donors = donorMaster.objects.filter(Status = 'No',Ornganization_id_id=orgid)
        for donor in ineligible_donors:
            try:
                last_donation = datetime.datetime.strptime(donor.Last_donation_date, "%Y-%m-%d").date()
                days_since_donation = (today - last_donation).days
                if days_since_donation >= 90:
                    donor.Status = 'Yes'
                    donor.save()
            except Exception as e:
                print(f"Error updating donor status:{str(e)}")
                pass
        datas = AreaMaster.objects.filter(Ornganization_id_id = orgid)
        return render(request,'USER_donor_search.html',{'datas':datas})
def USER_donor_search_db(request):
    if 'user' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            dt = userMaster.objects.get(User_id = request.session['user'])
            orgid = dt.Ornganization_id_id
            datas = AreaMaster.objects.filter(Ornganization_id_id = orgid)
            today = date.today()
            ineligible_donors = donorMaster.objects.filter(Status = 'No',Ornganization_id_id = orgid)
            for donor in ineligible_donors:
                try:
                    last_donation = datetime.datetime.strptime(donor.Last_donation_date,"%Y-%m-%d").date()
                    day_since_donation = (today-last_donation).days
                    if day_since_donation >= 90:
                        donor.Status = 'Yes'
                        donor.save()
                except Exception as e:
                    print(f"Error updating donor status: {str(e)}")
                    pass
            if request.POST.get('donor_name') != '':
                a = request.POST.get('donor_name')
                data = donorMaster.objects.filter(Donor_name = str(a)).filter(Ornganization_id_id = orgid).filter(Status = 'Yes')
                return render(request,'USER_donor_search.html',{'donors':data,'datas':datas})  
            if request.POST.get('mobile') != '':
                b = request.POST.get('mobile') 
                data = donorMaster.objects.filter(Phone_no = b).filter(Ornganization_id_id = orgid).filter(Status = 'Yes')
                return render(request,'USER_donor_search.html',{'donors':data,'datas':datas})
            if request.POST.get('blood_group') != '' and request.POST.get('area') != '':
                f = request.POST.get('blood_group')
                g = request.POST.get('area')
                data = donorMaster.objects.filter(Blood_group = f).filter(Area_id = int(g)).filter(Ornganization_id_id = orgid).filter(Status = 'Yes')
                return render(request,'USER_donor_search.html',{'donors':data,'datas':datas})
            if request.POST.get('blood_group') != '':
                c = request.POST.get('blood_group')
                data = donorMaster.objects.filter(Blood_group = c).filter(Ornganization_id_id = orgid).filter(Status = 'Yes')
                return render(request,'USER_donor_search.html',{'donors':data,'datas':datas})
            if request.POST.get('area') != '':
                d = request.POST.get('area')
                data = donorMaster.objects.filter(Area_id = d).filter(Ornganization_id_id = orgid).filter(Status = 'Yes')
                return render(request,'USER_donor_search.html',{'donors':data,'datas':datas})
              
        return render(request, 'USER_donor_search.html', {'datas': datas})
def USER2_donate_blood(request,Donor_id):
    if 'user' not in request.session:
        return redirect('/login')
    else:
        dt = userMaster.objects.get(User_id = request.session['user'])
        orid = dt.Ornganization_id_id
        datas1 = donorMaster.objects.get(Donor_id = Donor_id,Ornganization_id_id = orid)
        # print(datas1)
        datas = hospitalMaster.objects.filter(Ornganization_id_id = orid)
        return render(request,'USER2_donate_blood.html',{'donor':datas1,'datas':datas}) 
def USER2_donate_blood_db(request):
    if 'user' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            
            dd = donationDetails()
            dt1 = userMaster.objects.get(User_id = request.session['user'])
            dd.Ornganization_id_id = dt1.Ornganization_id_id
            dd.Captured_by = dt1
            hos_id = request.POST.get('hospital')
            dd.Hospital_id = hospitalMaster.objects.get(Hospital_id = hos_id)
            a = request.POST.get('donorId')
            b = donorMaster.objects.get(Donor_id = a)
            dd.Donor_id = b
            dd.Patient_bystander = request.POST.get('patientName')
            dd.Contact_no = request.POST.get('contactNo')
            dd.Purpose = request.POST.get('purpose')
            dd.Donation_date = request.POST.get('donationsdate')
            dd.Donation_status = request.POST.get('donationStatus')
            dd.Created_date = date.today().strftime("%Y-%m-%d")
            dd.Status = "Yes"
            dd.save()
            if dd.Donation_status == 'Donated':
                b.Last_donation_date = dd.Donation_date
                b.Status = 'No'
                b.save()
            messages.success(request,"Blood donated successfully")
            return redirect('/USER_donor_search')
def USER_donor_edit(request,Donor_id):
    if 'user' not in request.session:
        return redirect('/login')
    else:
        dt = userMaster.objects.get(User_id = request.session['user'])
        orid = dt.Ornganization_id_id
        datas1 = donorMaster.objects.get(Donor_id = Donor_id,Ornganization_id_id = orid)
        datas = AreaMaster.objects.filter(Ornganization_id_id=orid)
        return render(request,'USER_donor_edit.html',{'datas1':datas1,'datas':datas})
def USER_donor_edit_db(request):
    if 'user' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            donor_id = request.POST.get('donor_id')
            dt = userMaster.objects.get(User_id = request.session['user'])
            orid = dt.Ornganization_id_id
            dd = donorMaster.objects.get(Donor_id=donor_id)
            dd.Donor_name = request.POST.get('donor')
            dd.Phone_no = request.POST.get('phone')
            dd.Address = request.POST.get('address')
            area_id = request.POST.get('area')
            dd.Area = AreaMaster.objects.get(Area_code = area_id) 
            dd.save()
            messages.success(request,"Donor details edited successfully")
            return redirect('/USER_donor_search')
def USER_donor_report(request):
    if 'user' not in request.session:
        return redirect('/login')
    else:
        return render(request,'USER_donation_report.html')
def USER_donor_report_find(request):
    if 'user' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':
            fromdate = request.POST.get('fromdate')
            todate = request.POST.get('todate')
            
            request.session['fromdate'] = fromdate
            request.session['todate'] = todate
            
            
            dt = userMaster.objects.get(User_id = request.session['user'])
            orgid = dt.Ornganization_id_id
            donation_details = donationDetails.objects.filter(Donation_date__range=[fromdate, todate]).filter(Ornganization_id_id = orgid).filter(Donation_status = 'Donated')
            return render(request, 'USER_donation_report.html',{'donors':donation_details})
        return redirect('/USER_donor_report') 
def USER_export_donor_report(request):
    if 'user' not in request.session:
        return redirect('/login')
    else:
        fromdate = request.session.get('fromdate')
        todate = request.session.get('todate')
        
        dt = userMaster.objects.get(User_id=request.session['user'])
        orgid = dt.Ornganization_id_id
        donation_details = donationDetails.objects.filter(Donation_date__range=[fromdate, todate]).filter(Ornganization_id_id=orgid).filter(Donation_status='Donated')
        wb = Workbook()
        ws = wb.active
        ws.title = "Donation Report"
        headers = ['Donor Name', 'Donation Date', 'Hospital', 'Blood Group', 'Patient']
        for col,header in enumerate(headers,1):
            cell = ws.cell(row=1,column=col,value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row,donation in enumerate(donation_details,2):
            ws.cell(row=row,column=1,value=donation.Donor_id.Donor_name)
            ws.cell(row=row, column=2, value=str(donation.Donation_date))
            ws.cell(row=row, column=3, value=donation.Hospital_id.Hospital_name)
            ws.cell(row=row, column=4, value=donation.Donor_id.Blood_group)
            ws.cell(row=row, column=5, value=donation.Patient_bystander)
        
        for col in range(1,len(headers)+1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].auto_size = True
            
        response = HttpResponse(content_type='application/vandapplication/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Donor_Report.xlsx'
        wb.save(response)
        return response
def USER_view_donors(request):
    if 'user' not in request.session:
        return redirect('/login')
    else:
        dt = userMaster.objects.get(User_id = request.session['user'])
        orgid = dt.Ornganization_id_id
        datas = AreaMaster.objects.filter(Ornganization_id_id = orgid)
        return render(request,'USER_view_donors.html',{'datas':datas})
def USER_view_donors_db(request):
    if 'user' not in request.session:
        return redirect('/login')
    else:
        if request.method == 'POST':  
            dt = userMaster.objects.get(User_id = request.session['user'])
            orgid = dt.Ornganization_id_id
            datas = AreaMaster.objects.filter(Ornganization_id_id = orgid)
            if request.POST.get('donor_name') != '':
                a = request.POST.get('donor_name')
                data = donorMaster.objects.filter(Donor_name = str(a)).filter(Ornganization_id_id = orgid)
                return render(request,'USER_view_donors.html',{'donors':data,'datas':datas})  
            if request.POST.get('mobile') != '':
                b = request.POST.get('mobile') 
                data = donorMaster.objects.filter(Phone_no = b).filter(Ornganization_id_id = orgid)
                return render(request,'USER_view_donors.html',{'donors':data,'datas':datas})
            if request.POST.get('blood_group') != '' and request.POST.get('area') != '':
                f = request.POST.get('blood_group')
                g = request.POST.get('area')
                data = donorMaster.objects.filter(Blood_group = f).filter(Area_id = int(g)).filter(Ornganization_id_id = orgid)
                return render(request,'USER_view_donors.html',{'donors':data,'datas':datas})
            if request.POST.get('blood_group') != '':
                c = request.POST.get('blood_group')
                data = donorMaster.objects.filter(Blood_group = c).filter(Ornganization_id_id = orgid)
                return render(request,'USER_view_donors.html',{'donors':data,'datas':datas})
            if request.POST.get('area') != '':
                d = request.POST.get('area')
                data = donorMaster.objects.filter(Area_id = d).filter(Ornganization_id_id = orgid)
                return render(request,'USER_view_donors.html',{'donors':data,'datas':datas})
              
        return render(request, 'USER_view_donors.html', {'datas': datas})
            
        
        


            

        
    

    

    

